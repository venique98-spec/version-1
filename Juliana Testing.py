# ukids_scheduler_app.py
# uKids Scheduler — Google Sheets read/write version
#
# This version fixes the main issues in the previous code:
# - Uses gspread/service account access instead of public CSV links.
# - Reads private Google Sheets directly from your Streamlit secrets.
# - Writes the generated schedule back to the Google Sheet tab: "Roster first draft".
# - Clears and rewrites the output tab each time you click Generate Schedule.
# - Keeps the Excel download option.
# - Keeps the priority logic, campus logic, Brooklyn rule, director ignoring, and leader group blocking.
#
# Required Streamlit secrets:
# 1) GOOGLE_SHEET_ID = "your_google_sheet_id"
# 2) Your Google service account credentials.
#
# Supported credential formats:
# Option A:
# [gcp_service_account]
# type = "service_account"
# project_id = "..."
# private_key_id = "..."
# private_key = "-----BEGIN PRIVATE KEY-----\n...\n-----END PRIVATE KEY-----\n"
# client_email = "..."
# client_id = "..."
# auth_uri = "https://accounts.google.com/o/oauth2/auth"
# token_uri = "https://oauth2.googleapis.com/token"
# auth_provider_x509_cert_url = "https://www.googleapis.com/oauth2/v1/certs"
# client_x509_cert_url = "..."
#
# Option B:
# GOOGLE_SERVICE_ACCOUNT_JSON = "{...full service account json...}"
#
# IMPORTANT:
# Share the Google Sheet with the service account client_email as Editor.

import io
import re
import json
import base64
from collections import defaultdict, Counter
from datetime import datetime, date

import pandas as pd
import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# Fixed Google Sheet settings
# ──────────────────────────────────────────────────────────────────────────────
FIXED_GOOGLE_SHEET_ID = ""  # Optional fallback. Prefer st.secrets["GOOGLE_SHEET_ID"]
FIXED_RESPONSES_TAB = "Responses"
FIXED_SERVINGBASE_TAB = "ServingBase"
FIXED_MAPPING_TAB = "Mapping sheet"
FIXED_OUTPUT_TAB = "Roster first draft"

# ──────────────────────────────────────────────────────────────────────────────
# Page setup
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="uKids Scheduler", layout="wide")
st.title("uKids Scheduler")

st.markdown(
    """
    <style>
      .stApp { background: #000; color: #fff; }
      .stButton>button, .stDownloadButton>button { background:#444; color:#fff; }
      .stDataFrame { background:#111; }
      .stAlert { color:#111; }
    </style>
    """,
    unsafe_allow_html=True,
)

for logo_name in ["image(1).png", "image.png", "logo.png"]:
    try:
        with open(logo_name, "rb") as f:
            encoded = base64.b64encode(f.read()).decode()
        st.markdown(
            f"<div style='text-align:center'><img src='data:image/png;base64,{encoded}' width='520'></div>",
            unsafe_allow_html=True,
        )
        break
    except Exception:
        pass

# ──────────────────────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────────────────────
MONTH_ALIASES = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}
YES_SET = {"yes", "y", "true", "available"}
NO_SET = {"no", "n", "false", "not available"}

CAMPUS_ORDER = ["UC", "BRK", "TGB", "NEL", "POL"]
CAMPUS_LABELS = {
    "UC": "Unite City",
    "BRK": "Brooklyn",
    "TGB": "Tygerberg",
    "NEL": "Nelspruit",
    "POL": "Polokwane",
}
DIRECTOR_CODES = {"DIR", "D", "ND", "PD", "TD"}
ADMIN_VALUES = {"", "N/A", "NA", "NONE", "NAN", "-"}
PRIORITY_COLS = [
    "1A", "1B", "1C", "1D", "1E",
    "2A", "2B", "2C", "2D", "2E",
    "3A", "3B", "3C", "3D", "3E",
    "4A", "4B", "5",
]

# ──────────────────────────────────────────────────────────────────────────────
# General helpers
# ──────────────────────────────────────────────────────────────────────────────
def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s).lower()).strip()


def norm_name(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def clean_text(s) -> str:
    if pd.isna(s):
        return ""
    return str(s).replace("\u00A0", " ").strip()


def clean_code(s) -> str:
    return clean_text(s).upper().replace(" ", "")


def is_blank_or_na(value) -> bool:
    return clean_code(value) in ADMIN_VALUES


def to_int_capacity(value) -> int:
    if pd.isna(value):
        return 0
    s = str(value).strip()
    if not s:
        return 0
    try:
        return max(0, int(float(s)))
    except Exception:
        return 0


def excel_autofit(ws):
    for col_idx, column_cells in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
        start=1,
    ):
        max_len = 0
        for cell in column_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 80)


def df_to_sheet_values(df: pd.DataFrame):
    if df.empty:
        return [["No data"]]
    safe_df = df.fillna("").astype(str)
    return [list(safe_df.columns)] + safe_df.values.tolist()

# ──────────────────────────────────────────────────────────────────────────────
# Google Sheets via gspread
# ──────────────────────────────────────────────────────────────────────────────
def get_fixed_sheet_id() -> str:
    sheet_id = FIXED_GOOGLE_SHEET_ID.strip()
    if sheet_id:
        return sheet_id
    try:
        sheet_id = str(st.secrets.get("GOOGLE_SHEET_ID", "")).strip()
        if sheet_id:
            return sheet_id
    except Exception:
        pass
    st.error("Google Sheet ID is missing. Add GOOGLE_SHEET_ID to Streamlit secrets, or paste it into FIXED_GOOGLE_SHEET_ID in the code.")
    st.stop()


def extract_sheet_id(value: str) -> str:
    value = str(value or "").strip()
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", value)
    if match:
        return match.group(1)
    return value


def get_service_account_info() -> dict:
    try:
        if "gcp_service_account" in st.secrets:
            return dict(st.secrets["gcp_service_account"])
        if "GOOGLE_SERVICE_ACCOUNT_JSON" in st.secrets:
            raw = st.secrets["GOOGLE_SERVICE_ACCOUNT_JSON"]
            return json.loads(raw) if isinstance(raw, str) else dict(raw)
    except Exception as e:
        st.error(f"Could not read Google service account details from secrets: {e}")
        st.stop()

    st.error(
        "Google service account credentials are missing. Add [gcp_service_account] or GOOGLE_SERVICE_ACCOUNT_JSON to Streamlit secrets."
    )
    st.stop()


@st.cache_resource(ttl=300)
def get_gspread_client():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    info = get_service_account_info()
    try:
        creds = Credentials.from_service_account_info(info, scopes=scopes)
        return gspread.authorize(creds)
    except Exception as e:
        st.error(f"Could not authorise Google Sheets. Check your service account secrets. Error: {e}")
        st.stop()


def open_workbook():
    sheet_id = extract_sheet_id(get_fixed_sheet_id())
    client = get_gspread_client()
    try:
        return client.open_by_key(sheet_id)
    except Exception as e:
        st.error(
            "Could not open the Google Sheet. Make sure the sheet ID is correct and the sheet is shared with the service account as Editor. "
            f"Error: {e}"
        )
        st.stop()


@st.cache_data(ttl=60)
def read_google_sheet_tab_cached(sheet_id: str, worksheet_name: str) -> pd.DataFrame:
    client = get_gspread_client()
    sh = client.open_by_key(extract_sheet_id(sheet_id))
    try:
        ws = sh.worksheet(worksheet_name)
    except Exception as e:
        raise RuntimeError(f"Could not find tab '{worksheet_name}'. Error: {e}")

    values = ws.get_all_values()
    if not values or not any(any(cell.strip() for cell in row) for row in values):
        raise RuntimeError(f"The tab '{worksheet_name}' is empty or could not be read.")

    headers = [str(h).strip() for h in values[0]]
    rows = values[1:]
    width = len(headers)
    fixed_rows = []
    for row in rows:
        fixed_rows.append((row + [""] * width)[:width])

    df = pd.DataFrame(fixed_rows, columns=headers)
    df = df.replace("", pd.NA).dropna(how="all").fillna("")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def read_google_sheet_tab(sheet_id: str, worksheet_name: str) -> pd.DataFrame:
    return read_google_sheet_tab_cached(sheet_id, worksheet_name)


def get_or_create_worksheet(sh, title: str, rows: int = 1000, cols: int = 50):
    try:
        return sh.worksheet(title)
    except gspread.WorksheetNotFound:
        return sh.add_worksheet(title=title, rows=rows, cols=cols)


def write_output_to_google_sheet(schedule_df, assignment_df, unscheduled_df, detected_dates_df, few_yes_df, unmet_p1_df, unknown_codes_df, meta_rows):
    sh = open_workbook()
    ws = get_or_create_worksheet(sh, FIXED_OUTPUT_TAB, rows=1000, cols=80)
    ws.clear()

    output = []
    output.append(["uKids Scheduler Output"])
    output.extend(meta_rows)
    output.append([])

    sections = [
        ("Schedule", schedule_df),
        ("Assignment Summary", assignment_df),
        ("Unscheduled but Available", unscheduled_df),
        ("Detected Dates", detected_dates_df),
        ("Fewer than 2 Yes", few_yes_df),
        ("Unmet Priority 1", unmet_p1_df),
    ]
    if unknown_codes_df is not None and not unknown_codes_df.empty:
        sections.append(("Unknown Codes", unknown_codes_df))

    for title, df in sections:
        output.append([title])
        output.extend(df_to_sheet_values(df))
        output.append([])
        output.append([])

    max_cols = max(len(row) for row in output) if output else 1
    normalised = [row + [""] * (max_cols - len(row)) for row in output]

    required_rows = max(len(normalised), 100)
    required_cols = max(max_cols, 20)
    try:
        ws.resize(rows=required_rows, cols=required_cols)
    except Exception:
        pass

    ws.update(normalised, value_input_option="USER_ENTERED")

    try:
        ws.freeze(rows=1)
        ws.format("A1:Z1", {"textFormat": {"bold": True}})
    except Exception:
        pass

    return ws.url

# ──────────────────────────────────────────────────────────────────────────────
# Column detection
# ──────────────────────────────────────────────────────────────────────────────
def get_column_by_candidates(df: pd.DataFrame, candidates: list[str], required=True, label="column"):
    cols_l = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        key = cand.strip().lower()
        if key in cols_l:
            return cols_l[key]
    if required:
        st.error(f"Could not detect {label}. Expected one of: {', '.join(candidates)}")
        st.stop()
    return None


def detect_serving_girl_column(df: pd.DataFrame) -> str:
    return get_column_by_candidates(
        df,
        ["Serving Girl", "Serving Girl Name", "Name", "Name & Surname", "Name and Surname", "Full name", "Full names"],
        required=True,
        label="Serving Girl column",
    )


def detect_timestamp_column(df: pd.DataFrame):
    return get_column_by_candidates(
        df,
        ["timestamp", "time stamp", "submitted at", "submission time"],
        required=False,
        label="timestamp column",
    )


def detect_availability_month_column(df: pd.DataFrame):
    return get_column_by_candidates(
        df,
        ["Availability", "Availability month", "Month"],
        required=False,
        label="availability month column",
    )

# ──────────────────────────────────────────────────────────────────────────────
# Mapping sheet parsing
# ──────────────────────────────────────────────────────────────────────────────
def parse_mapping_sheet(mapping_df: pd.DataFrame):
    code_col = get_column_by_candidates(
        mapping_df,
        ["Shortened Name", "Short Name", "Code", "Short code", "Shortened code"],
        required=True,
        label="mapping code column",
    )
    display_col = get_column_by_candidates(
        mapping_df,
        ["Display Name", "Display", "Role", "Role Name"],
        required=True,
        label="mapping display name column",
    )

    mapping = {}
    for _, row in mapping_df.iterrows():
        code = clean_code(row.get(code_col, ""))
        if is_blank_or_na(code):
            continue
        display = clean_text(row.get(display_col, "")) or code
        capacities = {campus: to_int_capacity(row.get(campus, 0)) if campus in mapping_df.columns else 0 for campus in CAMPUS_ORDER}
        display_norm = normalize(display)
        is_director = code in DIRECTOR_CODES or "director" in display_norm
        is_leader = "leader" in display_norm or code.endswith("L") or code in {"BL", "PL", "EL", "SL", "L"}
        mapping[code] = {
            "code": code,
            "display": display,
            "capacities": capacities,
            "is_director": is_director,
            "is_leader": is_leader,
        }

    if not mapping:
        st.error("No valid codes were found in the Mapping sheet.")
        st.stop()
    return mapping

# ──────────────────────────────────────────────────────────────────────────────
# ServingBase parsing
# ──────────────────────────────────────────────────────────────────────────────
def parse_codes_from_cell(value):
    if is_blank_or_na(value):
        return []
    text = clean_text(value).upper()
    text = re.sub(r"[;,/|]+", " ", text)
    tokens = re.findall(r"[A-Z]+[0-9]*[A-Z]*|[A-Z][0-9]+[A-Z]+", text)
    cleaned = []
    for token in tokens:
        token = token.strip().upper()
        if token and token not in ADMIN_VALUES:
            cleaned.append(token)
    return cleaned


def priority_from_col(col_name: str) -> int | None:
    col = str(col_name).strip().upper()
    if col == "5":
        return 5
    match = re.match(r"^([1-4])[A-E]$", col)
    if match:
        return int(match.group(1))
    return None


def parse_serving_base(serving_df: pd.DataFrame, mapping: dict):
    director_col = get_column_by_candidates(serving_df, ["Director"], required=False, label="Director column")
    name_col = get_column_by_candidates(serving_df, ["Serving Girl", "Name", "Full name"], required=True, label="Serving Girl column")
    campus_col = get_column_by_candidates(serving_df, ["Primary Campus", "Campus"], required=True, label="Primary Campus column")
    position_col = get_column_by_candidates(serving_df, ["Position", "Code", "Role Code"], required=False, label="Position column")
    group_col = get_column_by_candidates(serving_df, ["Group"], required=False, label="Group column")

    upper_col_map = {str(c).strip().upper(): c for c in serving_df.columns}
    available_priority_cols = [upper_col_map[c] for c in PRIORITY_COLS if c in upper_col_map]
    if not available_priority_cols:
        st.error("No priority columns found in ServingBase. Expected columns like 1A, 1B, 2A, 3A, 4A, 4B, 5.")
        st.stop()

    people = {}
    ignored_directors = []
    unknown_codes = defaultdict(set)

    for _, row in serving_df.iterrows():
        name = clean_text(row.get(name_col, ""))
        person = norm_name(name)
        if not person:
            continue

        position_raw = clean_text(row.get(position_col, "")) if position_col else ""
        position_codes = set(parse_codes_from_cell(position_raw))
        if position_codes & DIRECTOR_CODES or "dir" in normalize(position_raw):
            ignored_directors.append(name)
            continue

        campus = clean_code(row.get(campus_col, ""))
        if campus not in CAMPUS_ORDER:
            continue

        group = clean_code(row.get(group_col, "")) if group_col else ""
        if group not in {"A", "B"}:
            group = ""

        priorities = defaultdict(set)
        all_codes = set()
        for col in available_priority_cols:
            priority = priority_from_col(col)
            if priority is None:
                continue
            for code in parse_codes_from_cell(row.get(col, "")):
                if code in DIRECTOR_CODES:
                    continue
                if code not in mapping:
                    unknown_codes[code].add(name)
                    continue
                if mapping[code]["is_director"]:
                    continue
                priorities[priority].add(code)
                all_codes.add(code)

        if not all_codes:
            continue

        people[person] = {
            "person": person,
            "display": name,
            "director": clean_text(row.get(director_col, "")) if director_col else "",
            "campus": campus,
            "group": group,
            "priorities": {p: sorted(codes) for p, codes in priorities.items()},
            "all_codes": all_codes,
        }

    if not people:
        st.error("No schedulable serving girls were found in ServingBase after directors and invalid rows were ignored.")
        st.stop()
    return people, ignored_directors, unknown_codes

# ──────────────────────────────────────────────────────────────────────────────
# Responses / availability parsing
# ──────────────────────────────────────────────────────────────────────────────
def extract_day_month_from_text(text: str):
    parts = re.findall(r"[0-9]{1,2}|[A-Za-z]{3,}", str(text))
    for i in range(len(parts) - 1):
        if parts[i].isdigit() and parts[i + 1].isalpha():
            day = int(parts[i])
            month = MONTH_ALIASES.get(parts[i + 1].lower()[:3])
            if month:
                return day, month
    return None


def infer_year_from_responses(responses_df: pd.DataFrame, service_month: int) -> int:
    month_col = detect_availability_month_column(responses_df)
    if month_col:
        values = [clean_text(v) for v in responses_df[month_col].dropna() if clean_text(v)]
        for raw in values:
            parsed = pd.to_datetime(raw, errors="coerce")
            if pd.notna(parsed):
                return int(parsed.year)
            match = re.search(r"(20\d{2})[-/](\d{1,2})", raw)
            if match:
                return int(match.group(1))

    submission_year = date.today().year
    submission_month = date.today().month
    timestamp_col = detect_timestamp_column(responses_df)
    if timestamp_col:
        ts = pd.to_datetime(responses_df[timestamp_col], errors="coerce").dropna()
        if not ts.empty:
            latest = ts.max()
            submission_year = int(latest.year)
            submission_month = int(latest.month)

    if service_month == 1 and submission_month in (11, 12):
        return submission_year + 1
    if service_month == 12 and submission_month in (1, 2):
        return submission_year - 1
    return submission_year


def detect_availability_date_columns(responses_df: pd.DataFrame):
    admin_norm = {normalize(x) for x in [
        "timestamp", "time stamp", "availability", "availability month", "month",
        "director", "serving girl", "serving girl name", "reason",
    ]}
    detected = []
    for col in responses_df.columns:
        if normalize(col) in admin_norm:
            continue
        parsed = extract_day_month_from_text(str(col))
        if not parsed:
            continue
        values = [clean_text(v).lower() for v in responses_df[col].dropna().tolist()]
        yes_no_count = sum(1 for v in values if v in YES_SET or v in NO_SET)
        if yes_no_count > 0 or len(values) == 0:
            detected.append(col)

    if not detected:
        st.error("No availability date columns were detected in Responses. Use headers like '7 June' or '21 June - Family Service'.")
        st.stop()
    return detected


def build_date_map_from_responses(responses_df: pd.DataFrame):
    date_cols = detect_availability_date_columns(responses_df)
    month_info = []
    for col in date_cols:
        parsed = extract_day_month_from_text(str(col))
        if parsed:
            day, month = parsed
            month_info.append((col, day, month))

    months = {m for _, _, m in month_info}
    if len(months) > 1:
        st.error(f"Multiple months detected in response date columns: {sorted(months)}. Please schedule one month at a time.")
        st.stop()

    service_month = list(months)[0]
    service_year = infer_year_from_responses(responses_df, service_month)
    date_map = {col: pd.Timestamp(datetime(service_year, month, day)).normalize() for col, day, month in month_info}
    service_dates = sorted(set(date_map.values()))
    sheet_name = f"{service_dates[0]:%B %Y}"
    return date_map, service_dates, sheet_name


def dedupe_latest_responses(responses_df: pd.DataFrame, person_col: str):
    df = responses_df.copy()
    df["_person"] = df[person_col].map(norm_name)
    df = df[df["_person"].astype(str).str.strip() != ""]
    timestamp_col = detect_timestamp_column(df)
    if timestamp_col:
        df["_ts"] = pd.to_datetime(df[timestamp_col], errors="coerce")
        df = df.sort_values("_ts")
        return df.groupby("_person", as_index=False).tail(1).drop(columns=["_ts"])
    return df.groupby("_person", as_index=False).tail(1)


def parse_availability(responses_df: pd.DataFrame, people: dict, date_map: dict):
    person_col = detect_serving_girl_column(responses_df)
    latest = dedupe_latest_responses(responses_df, person_col)
    availability = {person: {dt: False for dt in date_map.values()} for person in people.keys()}
    display_from_responses = {}
    yes_counts = Counter()
    responders_not_in_base = set()

    for _, row in latest.iterrows():
        person = norm_name(row.get(person_col, ""))
        if person not in people:
            if person:
                responders_not_in_base.add(clean_text(row.get(person_col, "")))
            continue
        display_from_responses[person] = clean_text(row.get(person_col, ""))
        for col, dt in date_map.items():
            ans = clean_text(row.get(col, "")).lower()
            is_yes = ans in YES_SET
            availability[person][dt] = is_yes
            if is_yes:
                yes_counts[person] += 1

    few_yes = sorted([p for p in people.keys() if yes_counts[p] < 2])
    return availability, few_yes, display_from_responses, sorted(responders_not_in_base)

# ──────────────────────────────────────────────────────────────────────────────
# Slot building
# ──────────────────────────────────────────────────────────────────────────────
def build_role_slots(mapping: dict, include_campuses: list[str]):
    slots = []
    for campus in include_campuses:
        for code, info in mapping.items():
            if info["is_director"]:
                continue
            capacity = info["capacities"].get(campus, 0)
            if capacity <= 0:
                continue
            for idx in range(1, capacity + 1):
                slot_id = f"{campus}|{code}|{idx}"
                slot_label = f"{campus} - {info['display']}" if capacity == 1 else f"{campus} - {info['display']} #{idx}"
                slots.append({
                    "slot_id": slot_id,
                    "campus": campus,
                    "code": code,
                    "display": info["display"],
                    "slot_label": slot_label,
                    "is_leader": info["is_leader"],
                    "sort_key": (CAMPUS_ORDER.index(campus), normalize(info["display"]), idx),
                })
    slots.sort(key=lambda x: x["sort_key"])
    return slots

# ──────────────────────────────────────────────────────────────────────────────
# Scheduling rules
# ──────────────────────────────────────────────────────────────────────────────
def person_priority_for_code(person_info: dict, code: str):
    for priority in [1, 2, 3, 4, 5]:
        if code in person_info["priorities"].get(priority, []):
            return priority
    return None


def group_allowed_for_role(person_info: dict, slot: dict, leader_group: str):
    group = person_info.get("group", "")
    if not slot["is_leader"]:
        return True
    if group not in {"A", "B"}:
        return True
    return group == leader_group


def campus_allowed_for_role(person_info: dict, slot: dict):
    person_campus = person_info["campus"]
    slot_campus = slot["campus"]
    if slot_campus == person_campus:
        return True
    if slot_campus == "BRK" and person_campus == "UC" and slot["code"] in person_info["all_codes"]:
        return True
    return False


def is_assigned_on_date(schedule_cells, person, service_date):
    return any(person in names for (_slot_id, d), names in schedule_cells.items() if d == service_date)


def can_assign(*, person, person_info, slot, service_date, schedule_cells, availability, assign_count, monthly_cap, leader_group):
    if assign_count.get(person, 0) >= monthly_cap:
        return False
    if is_assigned_on_date(schedule_cells, person, service_date):
        return False
    if not availability.get(person, {}).get(service_date, False):
        return False
    if slot["code"] not in person_info["all_codes"]:
        return False
    if not campus_allowed_for_role(person_info, slot):
        return False
    if not group_allowed_for_role(person_info, slot, leader_group):
        return False
    return True


def candidate_sort_key(person, person_info, priority, assign_count, availability, service_dates):
    yes_total = sum(1 for d in service_dates if availability.get(person, {}).get(d, False))
    group_rank = 0 if person_info.get("group") in {"A", "B"} else 1
    return (
        assign_count.get(person, 0),
        priority if priority is not None else 99,
        yes_total,
        group_rank,
        person_info["display"].lower(),
    )

# ──────────────────────────────────────────────────────────────────────────────
# Scheduling engine
# ──────────────────────────────────────────────────────────────────────────────
def generate_schedule(people, mapping, availability, service_dates, monthly_cap, leader_group, include_campuses):
    slots = build_role_slots(mapping, include_campuses)
    schedule_cells = {(slot["slot_id"], d): [] for slot in slots for d in service_dates}
    assign_count = Counter()
    slots_by_id = {s["slot_id"]: s for s in slots}

    people_order = sorted(
        people.keys(),
        key=lambda p: (
            sum(1 for d in service_dates if availability.get(p, {}).get(d, False)),
            people[p]["display"].lower(),
        ),
    )

    # Pass 1: try to give every person one first-priority assignment.
    for person in people_order:
        person_info = people[person]
        p1_codes = set(person_info["priorities"].get(1, []))
        if not p1_codes:
            continue
        assigned = False
        for service_date in service_dates:
            if assigned:
                break
            if not availability.get(person, {}).get(service_date, False):
                continue
            candidate_slots = [
                s for s in slots
                if s["code"] in p1_codes and not schedule_cells[(s["slot_id"], service_date)]
            ]
            candidate_slots.sort(key=lambda s: (0 if s["campus"] == person_info["campus"] else 1, s["sort_key"]))
            for slot in candidate_slots:
                if can_assign(
                    person=person,
                    person_info=person_info,
                    slot=slot,
                    service_date=service_date,
                    schedule_cells=schedule_cells,
                    availability=availability,
                    assign_count=assign_count,
                    monthly_cap=monthly_cap,
                    leader_group=leader_group,
                ):
                    schedule_cells[(slot["slot_id"], service_date)].append(person)
                    assign_count[person] += 1
                    assigned = True
                    break

    # Pass 2: fill all remaining slots by priority 1 to 5.
    for service_date in service_dates:
        for slot in slots:
            key = (slot["slot_id"], service_date)
            if schedule_cells[key]:
                continue
            candidates = []
            for person, person_info in people.items():
                priority = person_priority_for_code(person_info, slot["code"])
                if priority is None:
                    continue
                if not can_assign(
                    person=person,
                    person_info=person_info,
                    slot=slot,
                    service_date=service_date,
                    schedule_cells=schedule_cells,
                    availability=availability,
                    assign_count=assign_count,
                    monthly_cap=monthly_cap,
                    leader_group=leader_group,
                ):
                    continue
                candidates.append((person, priority))
            if candidates:
                candidates.sort(key=lambda t: candidate_sort_key(t[0], people[t[0]], t[1], assign_count, availability, service_dates))
                chosen = candidates[0][0]
                schedule_cells[key].append(chosen)
                assign_count[chosen] += 1

    served_p1 = set()
    for (slot_id, _d), assigned_people in schedule_cells.items():
        slot = slots_by_id[slot_id]
        for person in assigned_people:
            if slot["code"] in people[person]["priorities"].get(1, []):
                served_p1.add(person)
    unmet_p1 = sorted([p for p in people if people[p]["priorities"].get(1) and p not in served_p1])
    return schedule_cells, assign_count, slots, unmet_p1

# ──────────────────────────────────────────────────────────────────────────────
# Output tables
# ──────────────────────────────────────────────────────────────────────────────
def build_schedule_df(schedule_cells, slots, service_dates, people):
    rows = []
    for slot in slots:
        row = {"Position / Slot": slot["slot_label"]}
        for d in service_dates:
            assigned = schedule_cells.get((slot["slot_id"], d), [])
            row[d.strftime("%Y-%m-%d")] = ", ".join(people[p]["display"] for p in assigned)
        rows.append(row)
    return pd.DataFrame(rows)


def build_assignment_summary(schedule_cells, slots, people):
    slot_by_id = {s["slot_id"]: s for s in slots}
    per = defaultdict(list)
    for (slot_id, d), assigned_people in schedule_cells.items():
        slot = slot_by_id[slot_id]
        for p in assigned_people:
            per[p].append(f"{d.strftime('%Y-%m-%d')} — {slot['slot_label']}")
    rows = []
    for p, items in per.items():
        rows.append({
            "Person": people[p]["display"],
            "Campus": people[p]["campus"],
            "Group": people[p]["group"],
            "Assignments": len(items),
            "Locations & Dates": "; ".join(items),
        })
    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["Person", "Campus", "Group", "Assignments", "Locations & Dates"])
    return df.sort_values(["Assignments", "Person"], ascending=[False, True]).reset_index(drop=True)


def build_unscheduled_available(schedule_cells, service_dates, people, availability):
    scheduled_by_date = defaultdict(set)
    for (_slot_id, d), assigned_people in schedule_cells.items():
        for p in assigned_people:
            scheduled_by_date[d].add(p)
    max_len = 0
    per_date = {}
    for d in service_dates:
        names = []
        for p, info in people.items():
            if availability.get(p, {}).get(d, False) and p not in scheduled_by_date[d]:
                names.append((CAMPUS_ORDER.index(info["campus"]), info["display"], info["campus"], info["group"]))
        names.sort(key=lambda x: (x[0], x[1]))
        per_date[d] = names
        max_len = max(max_len, len(names))
    data = {}
    for d in service_dates:
        date_label = d.strftime("%Y-%m-%d")
        data[date_label] = [x[1] for x in per_date[d]] + [""] * (max_len - len(per_date[d]))
        data[f"{date_label} campus"] = [x[2] for x in per_date[d]] + [""] * (max_len - len(per_date[d]))
        data[f"{date_label} group"] = [x[3] for x in per_date[d]] + [""] * (max_len - len(per_date[d]))
    return pd.DataFrame(data)


def build_detected_dates_df(date_map):
    return pd.DataFrame({
        "Response column": list(date_map.keys()),
        "Service date": [d.strftime("%Y-%m-%d") for d in date_map.values()],
    })


def build_unknown_codes_df(unknown_codes):
    rows = []
    for code, names in sorted(unknown_codes.items()):
        rows.append({"Unknown code": code, "Found for": ", ".join(sorted(names))})
    return pd.DataFrame(rows)


def build_not_in_base_df(responders_not_in_base):
    return pd.DataFrame({"Responder not found in ServingBase": responders_not_in_base})

# ──────────────────────────────────────────────────────────────────────────────
# UI
# ──────────────────────────────────────────────────────────────────────────────
st.subheader("1) Settings")
cap_options = [1, 2, 3, 4, 5, 6]
selected_cap = st.selectbox("Maximum number of times a girl may serve this month", options=cap_options, index=1)
leader_group = st.radio(
    "Which group serves as leaders this month?",
    options=["A", "B"],
    horizontal=True,
    help="The other group is blocked from leader roles for this schedule.",
)
include_campuses = st.multiselect(
    "Campuses to schedule",
    options=CAMPUS_ORDER,
    default=CAMPUS_ORDER,
    format_func=lambda x: f"{x} - {CAMPUS_LABELS.get(x, x)}",
)

st.caption(f"• Source tabs: {FIXED_RESPONSES_TAB}, {FIXED_SERVINGBASE_TAB}, {FIXED_MAPPING_TAB}.")
st.caption(f"• Output tab: {FIXED_OUTPUT_TAB}.")
st.caption("• Brooklyn is filled by UC people only where their own priority codes contain the Brooklyn role code.")
st.caption("• Director rows/codes are ignored for now.")

if st.button("Generate Schedule", type="primary"):
    if not include_campuses:
        st.error("Please select at least one campus to schedule.")
        st.stop()

    sheet_id = get_fixed_sheet_id()

    try:
        responses_df = read_google_sheet_tab(sheet_id, FIXED_RESPONSES_TAB)
        serving_df = read_google_sheet_tab(sheet_id, FIXED_SERVINGBASE_TAB)
        mapping_df = read_google_sheet_tab(sheet_id, FIXED_MAPPING_TAB)
    except Exception as e:
        st.error(str(e))
        st.stop()

    mapping = parse_mapping_sheet(mapping_df)
    people, ignored_directors, unknown_codes = parse_serving_base(serving_df, mapping)
    date_map, service_dates, sheet_name = build_date_map_from_responses(responses_df)
    availability, few_yes, display_from_responses, responders_not_in_base = parse_availability(responses_df, people, date_map)

    for person, display in display_from_responses.items():
        if person in people and display:
            people[person]["display"] = display

    schedule_cells, assign_count, slots, unmet_p1 = generate_schedule(
        people=people,
        mapping=mapping,
        availability=availability,
        service_dates=service_dates,
        monthly_cap=int(selected_cap),
        leader_group=leader_group,
        include_campuses=include_campuses,
    )

    schedule_df = build_schedule_df(schedule_cells, slots, service_dates, people)
    assignment_df = build_assignment_summary(schedule_cells, slots, people)
    unscheduled_df = build_unscheduled_available(schedule_cells, service_dates, people, availability)
    detected_dates_df = build_detected_dates_df(date_map)
    unknown_codes_df = build_unknown_codes_df(unknown_codes)

    few_yes_df = pd.DataFrame({
        "Person": [people[p]["display"] for p in few_yes if p in people],
        "Campus": [people[p]["campus"] for p in few_yes if p in people],
        "Group": [people[p]["group"] for p in few_yes if p in people],
    })

    if unmet_p1:
        unmet_p1_df = pd.DataFrame({
            "Person": [people[p]["display"] for p in unmet_p1],
            "Campus": [people[p]["campus"] for p in unmet_p1],
            "Group": [people[p]["group"] for p in unmet_p1],
        })
    else:
        unmet_p1_df = pd.DataFrame(columns=["Person", "Campus", "Group"])

    total_slots = len(slots) * len(service_dates)
    filled_slots = sum(1 for names in schedule_cells.values() if names)
    fill_rate = filled_slots / total_slots if total_slots else 0

    meta_rows = [
        ["Generated at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Schedule month", sheet_name],
        ["Filled slots", f"{filled_slots} / {total_slots}"],
        ["Fill rate", f"{fill_rate:.1%}"],
        ["Monthly cap", str(selected_cap)],
        ["Leader group", leader_group],
        ["Campuses", ", ".join(include_campuses)],
        ["Ignored director rows", str(len(ignored_directors))],
        ["Responders not found in ServingBase", str(len(responders_not_in_base))],
    ]

    try:
        output_url = write_output_to_google_sheet(
            schedule_df=schedule_df,
            assignment_df=assignment_df,
            unscheduled_df=unscheduled_df,
            detected_dates_df=detected_dates_df,
            few_yes_df=few_yes_df,
            unmet_p1_df=unmet_p1_df,
            unknown_codes_df=unknown_codes_df,
            meta_rows=meta_rows,
        )
        st.success(f"Schedule generated for **{sheet_name}** and written to **{FIXED_OUTPUT_TAB}**.")
        st.markdown(f"[Open the output tab]({output_url})")
    except Exception as e:
        st.error(f"The schedule was generated, but could not be written to Google Sheets. Error: {e}")

    st.write(
        f"Filled slots: **{filled_slots} / {total_slots}** "
        f"(Fill rate: **{fill_rate:.1%}**) • Monthly cap: **{selected_cap}** • Leader group: **{leader_group}**"
    )

    st.subheader("Detected availability dates")
    st.dataframe(detected_dates_df, use_container_width=True)

    if ignored_directors:
        st.info(f"Ignored director rows: {len(ignored_directors)}")

    if responders_not_in_base:
        st.warning("Some responses were submitted by names that were not found in ServingBase.")
        st.dataframe(build_not_in_base_df(responders_not_in_base), use_container_width=True)

    if not unknown_codes_df.empty:
        st.warning("Some ServingBase codes were not found in the Mapping sheet and were ignored.")
        st.dataframe(unknown_codes_df, use_container_width=True)

    st.subheader("Schedule")
    st.dataframe(schedule_df, use_container_width=True)

    st.subheader("Assignment Summary")
    st.dataframe(assignment_df, use_container_width=True)

    st.subheader("People with fewer than 2 Yes dates - info")
    st.dataframe(few_yes_df, use_container_width=True)

    st.subheader("Unmet Priority-1 - info")
    st.dataframe(unmet_p1_df, use_container_width=True)

    st.subheader("Unscheduled but Available")
    st.dataframe(unscheduled_df, use_container_width=True)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    def write_df(ws_name, df):
        ws = wb.create_sheet(ws_name[:31])
        if df.empty:
            ws.append(["No data"])
        else:
            ws.append(list(df.columns))
            for _, row in df.iterrows():
                ws.append([row.get(col, "") for col in df.columns])
        excel_autofit(ws)

    write_df(sheet_name, schedule_df)
    write_df("Assignment Summary", assignment_df)
    write_df("Detected Dates", detected_dates_df)
    write_df("Fewer than 2 Yes", few_yes_df)
    write_df("Unmet Priority 1", unmet_p1_df)
    write_df("Unscheduled", unscheduled_df)
    if not unknown_codes_df.empty:
        write_df("Unknown Codes", unknown_codes_df)
    if responders_not_in_base:
        write_df("Names Not In Base", build_not_in_base_df(responders_not_in_base))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    st.download_button(
        "Download Excel (.xlsx)",
        data=buf,
        file_name=f"uKids_schedule_{sheet_name.replace(' ', '_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.info("Choose the settings, then click **Generate Schedule**.")
