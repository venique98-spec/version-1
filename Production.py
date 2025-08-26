# ukids_scheduler_app.py
# uKids Scheduler — dynamic roles & capacities from spreadsheet headers + strict preferred-fill for Ages 9–11
# - Positions & capacities come from row-1 headers of the Positions CSV.
#   Use "(xN)" / "[xN]" / "xN" suffix to declare capacity >1, e.g. "Info (x4)".
#   Headers without a suffix default to capacity=1.
# - Robust CSV ingestion
# - Latest availability per person (by Timestamp)
# - Names matched case/space-insensitively across files (display uses Positions name)
# - Only people present in BOTH files can be scheduled
# - Preferences: 0=not allowed; 1=must serve once; 2/3/4=can serve (prefer 2>3>4)
# - Directors (D) only eligible for priority==1 roles
# - Exclude plain "Entrance greeter" (but not other greeter names you add)
# - One person per slot/date
# - P1 pre-pass + "Unmet Priority-1" report
# - STRICT preferred-fill order each Sunday: Age9 L, Age9 C, Age10 L, Age10 C, Age11 L, Age11 C
# - "Helping ninja and check in leader" requires a uKids leader
# - Excel export (Schedule, Assignment Summary w/ details, Fewer than 2 Yes, Unmet P1)

import io
import re
import base64
from collections import defaultdict, Counter
from datetime import datetime, date

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="uKids Scheduler", layout="wide")
st.title("uKids Scheduler")

st.markdown(
    """
    <style>
      .stApp { background: #000; color: #fff; }
      .stButton>button, .stDownloadButton>button { background:#444; color:#fff; }
      .stDataFrame { background:#111; }
      .stAlert { color:#111; }
    </style>
    """,
    unsafe_allow_html=True,
)

# Optional logo (ignore if missing)
for logo_name in ["image(1).png", "image.png", "logo.png"]:
    try:
        with open(logo_name, "rb") as img_file:
            encoded = base64.b64encode(img_file.read()).decode()
            st.markdown(
                f"<div style='text-align:center'><img src='data:image/png;base64,{encoded}' width='520'></div>",
                unsafe_allow_html=True,
            )
            break
    except Exception:
        pass

# ──────────────────────────────────────────────────────────────────────────────
# Constants & helpers
# ──────────────────────────────────────────────────────────────────────────────
MONTH_ALIASES = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}
YES_SET = {"yes", "y", "true", "available"}

# Roles to exclude entirely (even if present in the Positions CSV)
EXCLUDED_ROLES = {"entrance greeter"}  # removed per your earlier request

# Special roles requiring a uKids leader (normalize()d names)
REQUIRES_LEADER = {"helping ninja and check in leader"}

# Preferred-fill roles in strict order (base labels, case/space-insensitive match)
PREFERRED_FILL_ORDER = [
    "age 9 leader",
    "age 9 classroom",
    "age 10 leader",
    "age 10 classroom",
    "age 11 leader",
    "age 11 classroom",
]
PREFERRED_INDEX = {r: i for i, r in enumerate(PREFERRED_FILL_ORDER)}

def normalize(s: str) -> str:
    """Generic normalizer for roles/headers."""
    return re.sub(r"[^a-z0-9]+", " ", str(s).lower()).strip()

def norm_name(s: str) -> str:
    """Name normalizer: lower-case and collapse internal spaces."""
    return re.sub(r"\s+", " ", str(s).strip().lower())

# Parse header like "Info (x4)" / "Info [x4]" / "Info x4" / "Info [4]" → ("Info", 4)
CAP_PATTERNS = [
    re.compile(r"^(?P<base>.*?)[\s\-]*\(\s*x?\s*(?P<n>\d+)\s*\)\s*$", re.IGNORECASE),
    re.compile(r"^(?P<base>.*?)[\s\-]*\[\s*x?\s*(?P<n>\d+)\s*\]\s*$", re.IGNORECASE),
    re.compile(r"^(?P<base>.*?)[\s\-]*x\s*(?P<n>\d+)\s*$", re.IGNORECASE),
]

def parse_role_and_capacity(header: str):
    """Return (base_label, capacity:int). Defaults to (header, 1) if no suffix."""
    s = str(header).strip()
    for pat in CAP_PATTERNS:
        m = pat.match(s)
        if m:
            base = m.group("base").strip()
            n = int(m.group("n"))
            return (base if base else header, max(1, n))
    return (header, 1)

def strip_capacity_tag(role_label: str) -> str:
    """Remove any capacity marker from a header for matching/eligibility."""
    base, _ = parse_role_and_capacity(role_label)
    return base

def read_csv_robust(uploaded_file, label_for_error):
    """Read a Streamlit UploadedFile into a DataFrame, trying multiple encodings and separators."""
    raw = uploaded_file.getvalue()
    encodings = ["utf-8", "utf-8-sig", "cp1252", "iso-8859-1"]
    seps = [None, ",", ";", "\t", "|"]
    last_err = None
    for enc in encodings:
        for sep in seps:
            try:
                df = pd.read_csv(io.BytesIO(raw), encoding=enc, engine="python", sep=sep)
                if df.shape[1] == 0:
                    raise ValueError("Parsed 0 columns.")
                return df
            except Exception as e:
                last_err = f"{type(e).__name__}: {e}"
                continue
    st.error(
        f"Could not read {label_for_error} CSV. Last error: {last_err}. "
        "Try re-exporting as CSV (UTF-8) or remove unusual characters in headers."
    )
    st.stop()

def detect_name_column(df: pd.DataFrame, fallback_first: bool = True) -> str:
    candidates = [
        "What is your name AND surname?",
        "What is your name and surname?",
        "Name",
        "Full name",
        "Full names",
    ]
    cols_l = {str(c).strip().lower(): c for c in df.columns}
    for c in candidates:
        key = c.strip().lower()
        if key in cols_l:
            return cols_l[key]
    for c in df.columns:
        if isinstance(c, str) and "name" in c.lower():
            return c
    if fallback_first:
        return df.columns[0]
    raise ValueError("Could not detect a 'name' column.")

def is_priority_col(series: pd.Series) -> bool:
    vals = pd.to_numeric(series, errors="coerce").dropna()
    if len(vals) == 0:
        return False
    return (vals.min() >= 0) and (vals.max() <= 5)

def parse_month_and_dates_from_headers(responses_df: pd.DataFrame):
    avail_cols = [c for c in responses_df.columns if isinstance(c, str) and c.strip().lower().startswith("are you available")]
    if not avail_cols:
        avail_cols = [
            c for c in responses_df.columns
            if isinstance(c, str)
            and re.search(r"\b\d{1,2}\b", c.lower())
            and any(m in c.lower() for m in MONTH_ALIASES)
        ]
    if not avail_cols:
        raise ValueError("No availability columns found. Expect headings like 'Are you available 7 September?'")

    info = []
    for c in avail_cols:
        low = c.lower()
        mname = None
        for alias in MONTH_ALIASES:
            if alias in low:
                mname = alias
                break
        day_m = re.search(r"\b(\d{1,2})\b", low)
        if mname and day_m:
            info.append((c, MONTH_ALIASES[mname], int(day_m.group(1))))
    if not info:
        raise ValueError("Could not parse day/month from availability headers.")

    months = {m for _, m, _ in info}
    if len(months) > 1:
        raise ValueError(f"Multiple months detected in availability headers: {sorted(months)}. Upload one month at a time.")
    month = months.pop()

    if "Timestamp" in responses_df.columns:
        years = pd.to_datetime(responses_df["Timestamp"], errors="coerce").dt.year.dropna().astype(int)
        year = int(years.mode().iloc[0]) if not years.empty else date.today().year
    else:
        year = date.today().year

    date_map = {c: pd.Timestamp(datetime(year, month, d)).normalize() for c, _, d in info}
    service_dates = sorted(set(date_map.values()))
    sheet_name = f"{pd.Timestamp(year=year, month=month, day=1):%B %Y}"
    return year, month, date_map, service_dates, sheet_name

# ──────────────────────────────────────────────────────────────────────────────
# Data shaping
# ──────────────────────────────────────────────────────────────────────────────
def build_display_name_map(positions_df: pd.DataFrame, name_col: str):
    """Prefer display casing from Positions file."""
    disp = {}
    for _, r in positions_df.iterrows():
        raw = str(r[name_col]).strip()
        if raw:
            disp.setdefault(norm_name(raw), raw)
    return disp

def build_long_df(people_df: pd.DataFrame, name_col: str, role_cols, codes_col: str = None):
    """
    Returns:
      - long_df with columns [person_norm, role_header, priority]
      - role_codes flags per person_norm: has_D / has_BL / has_PL / has_EL / has_SL + raw_codes
    """
    records = []
    role_codes = {}
    for _, r in people_df.iterrows():
        display_name = str(r[name_col]).strip()
        if not display_name or display_name.lower() == "nan":
            continue
        person = norm_name(display_name)  # normalized key

        # parse code flags (2nd column)
        flags = {"has_D": False, "has_BL": False, "has_PL": False, "has_EL": False, "has_SL": False, "raw": ""}
        if codes_col and codes_col in people_df.columns:
            raw = str(r.get(codes_col, "") or "")
            flags["raw"] = raw
            toks = re.findall(r"[A-Za-z]+", raw.upper())
            for t in toks:
                if t == "D": flags["has_D"] = True
                elif t == "BL": flags["has_BL"] = True
                elif t == "PL": flags["has_PL"] = True
                elif t == "EL": flags["has_EL"] = True
                elif t == "SL": flags["has_SL"] = True
        role_codes[person] = flags

        # preferences per role (use header as-is; we'll strip capacity when matching)
        for role_hdr in role_cols:
            pr = pd.to_numeric(r[role_hdr], errors="coerce")
            if pd.isna(pr):
                continue
            pr = int(round(pr))
            if pr >= 1:
                # Directors only eligible for priority==1 roles
                if flags["has_D"] and pr != 1:
                    continue
                records.append({"person": person, "role": role_hdr, "priority": pr})

    return pd.DataFrame(records), role_codes

def dedupe_latest_by_key(df: pd.DataFrame, key_series: pd.Series) -> pd.DataFrame:
    """Keep only the most recent row per normalized key using 'Timestamp' when present."""
    key_norm = key_series.map(norm_name)
    df2 = df.assign(_key=key_norm)
    if "Timestamp" in df2.columns:
        ts = pd.to_datetime(df2["Timestamp"], errors="coerce")
        df2 = df2.assign(_ts=ts).sort_values("_ts")
        latest = df2.groupby("_key", as_index=False).tail(1).drop(columns=["_ts"])
        return latest
    # Fallback: keep last seen per key
    return df2.groupby("_key", as_index=False).tail(1)

def parse_availability(responses_df: pd.DataFrame, name_col_resp: str, date_map):
    # Only latest response per normalized name
    responses_latest = dedupe_latest_by_key(responses_df, responses_df[name_col_resp])

    availability = {}
    yes_counts = Counter()
    display_from_responses = {}

    for _, row in responses_latest.iterrows():
        disp_name = str(row.get(name_col_resp, "")).strip()
        key = str(row.get("_key", "")).strip()
        if not key or key.lower() == "nan":
            continue
        display_from_responses[key] = disp_name

        availability.setdefault(key, {})
        for col, dt in date_map.items():
            ans = str(row.get(col, "")).strip().lower()
            is_yes = ans in YES_SET
            availability[key][dt] = is_yes
            if is_yes:
                yes_counts[key] += 1

    few_yes = sorted([n for n, c in yes_counts.items() if c < 2])
    service_dates = sorted(set(date_map.values()))
    return availability, service_dates, few_yes, display_from_responses

# Build slot plan from headers (strip capacity markers); combine duplicates by max capacity
def build_slot_plan_dynamic(all_role_headers):
    """
    all_role_headers: list of headers from positions CSV (3rd column onward)
    Returns: dict {base_role_label: capacity}
    """
    slot_plan = {}
    excluded = {normalize(x) for x in EXCLUDED_ROLES}
    for hdr in all_role_headers:
        base_label, cap = parse_role_and_capacity(hdr)
        if normalize(base_label) in excluded:
            continue
        # Take max if the same base label appears multiple times
        slot_plan[base_label] = max(cap, slot_plan.get(base_label, 0))
    return slot_plan

def expand_roles_to_slots(slot_plan):
    """Expand capacities into row labels, e.g., 'Info' x4 → 'Info #1'..'#4'."""
    slot_rows = []
    slot_index = {}
    for role, n in slot_plan.items():
        if n <= 0:
            continue
        if n == 1:
            lab = role
            slot_rows.append(lab)
            slot_index[lab] = role
        else:
            for i in range(1, n + 1):
                lab = f"{role} #{i}"
                slot_rows.append(lab)
                slot_index[lab] = role
    return slot_rows, slot_index

def build_eligibility(long_df: pd.DataFrame):
    """Return {person_norm: set(role_headers)} for pr>=1 (headers kept as-is)."""
    elig = defaultdict(set)
    for _, r in long_df.iterrows():
        elig[str(r["person"]).strip()].add(str(r["role"]).strip())
    return elig

def build_priority_lookup(long_df: pd.DataFrame):
    """Return {(person_norm, base_role_norm): priority}, using stripped base role."""
    lut = {}
    for _, r in long_df.iterrows():
        base = strip_capacity_tag(str(r["role"]))
        lut[(str(r["person"]).strip(), normalize(base))] = int(r["priority"])
    return lut

def is_ukids_leader(flags: dict) -> bool:
    return bool(flags.get("has_BL") or flags.get("has_PL") or flags.get("has_EL") or flags.get("has_SL") or flags.get("has_D"))

def base_max_for_person(flags: dict) -> int:
    # Monthly caps: Director=1, others=2
    return 1 if flags.get("has_D", False) else 2

def role_allowed_for_person(eligibility, person_norm, base_role):
    """Check if 'base_role' (stripped) is allowed for 'person_norm' (pr>=1)."""
    nb = normalize(strip_capacity_tag(base_role))
    for er in eligibility.get(person_norm, set()):
        if normalize(strip_capacity_tag(er)) == nb:
            return True
    return False

def pref_rank(val):
    """Lower = better. Prefer 2, then 3, then 4, then 1 (after P1 handled)."""
    if val == 2: return 0
    if val == 3: return 1
    if val == 4: return 2
    if val == 1: return 3
    return 9

def get_priority_for(lookup, person_norm, role_name):
    return lookup.get((person_norm, normalize(strip_capacity_tag(role_name))))

def compute_p1_roles_by_person(long_df, allowed_roles_set):
    """Return {person_norm: set(base_roles with priority==1)}, filtered to scheduled roles."""
    p1 = defaultdict(set)
    allowed_norm = {normalize(strip_capacity_tag(r)) for r in allowed_roles_set}
    for _, r in long_df.iterrows():
        base = strip_capacity_tag(str(r["role"]).strip())
        if normalize(base) not in allowed_norm:
            continue
        if int(r["priority"]) == 1:
            p1[str(r["person"]).strip()].add(base)
    return p1

def served_in_priority_one(schedule_cells, p1_roles_by_person, slot_to_role):
    """Return set of person_norm who got any slot whose base role is in their P1 list."""
    served = set()
    for (row_name, _d), names in schedule_cells.items():
        base_role = slot_to_role.get(row_name, row_name)
        for nm_norm in names:
            if base_role in p1_roles_by_person.get(nm_norm, set()):
                served.add(nm_norm)
    return served

# ──────────────────────────────────────────────────────────────────────────────
# Scheduling
# ──────────────────────────────────────────────────────────────────────────────
def main_pass_schedule(long_df, availability, service_dates, role_codes, all_role_headers):
    # Build slots dynamically (exclude removed roles); headers may contain (xN)
    slot_plan = build_slot_plan_dynamic(all_role_headers)
    slot_rows, slot_to_role = expand_roles_to_slots(slot_plan)

    eligibility = build_eligibility(long_df)
    priority_lut = build_priority_lookup(long_df)

    # Only schedule people present in BOTH data sources (keys are normalized)
    people = sorted(set(eligibility.keys()) & set(availability.keys()))

    # {(row_label, date): [person_norm]}
    schedule_cells = {(slot, d): [] for slot in slot_rows for d in service_dates}
    assign_count = defaultdict(int)

    def slot_sort_key(s):
        """Order slots so preferred-fill roles come first IN EXACT ORDER, then other leaders, classrooms, others."""
        base_role = slot_to_role[s]
        n = normalize(base_role)
        if n in PREFERRED_INDEX:
            return (0, PREFERRED_INDEX[n], s.lower())  # strict order you requested
        s_low = s.lower()
        if "leader" in s_low:
            return (1, s_low)
        if "classroom" in s_low:
            return (2, s_low)
        return (3, s_low)

    slot_rows_sorted = sorted(slot_rows, key=slot_sort_key)

    # P1 pre-pass: try to give each P1-eligible person exactly one P1 slot
    p1_roles_by_person = compute_p1_roles_by_person(long_df, allowed_roles_set=slot_plan.keys())
    avail_count = {p: sum(1 for d in service_dates if availability.get(p, {}).get(d, False)) for p in people}
    p1_people_order = sorted([p for p in people if p1_roles_by_person.get(p)], key=lambda p: (avail_count.get(p, 0), p))

    for p in p1_people_order:
        flags = role_codes.get(p, {})
        cap = base_max_for_person(flags)
        if assign_count[p] >= cap:
            continue
        got_one = False
        for d in service_dates:
            if not availability.get(p, {}).get(d, False):
                continue
            # already assigned that day?
            if any(p in names for (rn, dd), names in schedule_cells.items() if dd == d):
                continue
            # find a free P1 slot (preferred-fill slots appear earlier in slot_rows_sorted)
            for slot_row in slot_rows_sorted:
                base_role = slot_to_role[slot_row]
                if base_role not in p1_roles_by_person[p]:
                    continue
                if not role_allowed_for_person(eligibility, p, base_role):
                    continue
                # leader-only gate
                if normalize(base_role) in REQUIRES_LEADER and not is_ukids_leader(flags):
                    continue
                # free?
                if len(schedule_cells[(slot_row, d)]) == 0:
                    schedule_cells[(slot_row, d)].append(p)
                    assign_count[p] += 1
                    got_one = True
                    break
            if got_one:
                break
        # if not got_one, they’ll appear in "Unmet P1" later

    # General fill (preferred-fill slots still come first)
    for d in service_dates:
        # set of people already on this date
        assigned_today = set(nm for (rn, dd), names in schedule_cells.items() if dd == d for nm in names)
        for slot_row in slot_rows_sorted:
            base_role = slot_to_role[slot_row]
            # respect single-person-per-slot
            if len(schedule_cells[(slot_row, d)]) >= 1:
                continue

            # collect candidates
            cands = []
            for p in people:
                flags = role_codes.get(p, {})
                if assign_count[p] >= base_max_for_person(flags):
                    continue
                if p in assigned_today:
                    continue
                if not availability.get(p, {}).get(d, False):
                    continue
                if not role_allowed_for_person(eligibility, p, base_role):
                    continue
                if normalize(base_role) in REQUIRES_LEADER and not is_ukids_leader(flags):
                    continue
                # compute preference for sorting (2>3>4>1)
                pr = get_priority_for(priority_lut, p, base_role)
                cands.append((p, pr, flags))

            if cands:
                # sort by (fewest assignments so far, best preference rank, then name)
                cands.sort(key=lambda t: (assign_count[t[0]], (0 if t[1]==2 else 1 if t[1]==3 else 2 if t[1]==4 else 3 if t[1]==1 else 9), t[0]))
                chosen = cands[0][0]
                schedule_cells[(slot_row, d)].append(chosen)
                assign_count[chosen] += 1
                assigned_today.add(chosen)

    return schedule_cells, assign_count, slot_rows, slot_to_role, eligibility, people, p1_roles_by_person

# ──────────────────────────────────────────────────────────────────────────────
# Output helpers
# ──────────────────────────────────────────────────────────────────────────────
def excel_autofit(ws):
    for col_idx, column_cells in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1
    ):
        max_len = 0
        for cell in column_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 80)

def build_schedule_df(schedule_cells, slot_rows, service_dates, name_display_map):
    cols = [d.strftime("%Y-%m-%d") for d in service_dates]
    df = pd.DataFrame(index=slot_rows, columns=cols)
    for (slot_row, d), names in schedule_cells.items():
        # map normalized names to display names (fallback to the key if missing)
        disp = [name_display_map.get(nm, nm) for nm in names]
        df.loc[slot_row, d.strftime("%Y-%m-%d")] = ", ".join(disp)
    return df.fillna("")

def build_person_assignment_details(schedule_cells, name_display_map):
    """Return dict: display_name -> 'YYYY-MM-DD — Slot; ...' (sorted by date)."""
    per = defaultdict(list)
    for (slot_row, d), names in schedule_cells.items():
        for nm_norm in names:
            per[nm_norm].append((d, slot_row))
    details = {}
    for nm_norm, items in per.items():
        items_sorted = sorted(items, key=lambda x: x[0])
        disp_name = name_display_map.get(nm_norm, nm_norm)
        details[disp_name] = "; ".join([f"{dt.strftime('%Y-%m-%d')} — {slot}" for dt, slot in items_sorted])
    return details

# ──────────────────────────────────────────────────────────────────────────────
# UI
# ──────────────────────────────────────────────────────────────────────────────
st.subheader("1) Upload files (CSV — any filename)")
c1, c2 = st.columns(2)
with c1:
    positions_file = st.file_uploader("Serving positions (CSV)", type=["csv"], key="positions_csv_any")
with c2:
    responses_file = st.file_uploader("Availability responses (CSV)", type=["csv"], key="responses_csv_any")

st.caption("• Positions CSV: first col = volunteer names; second col = role codes (e.g., D/BL/PL/EL/SL); other cols = roles with values 0–5 (0 = not allowed, 1 = must serve once).")
st.caption("• For multi-slot roles, add '(xN)' to the header, e.g., 'Info (x4)'.")
st.caption("• Responses CSV: includes a name column and availability columns like 'Are you available 7 September?' plus a Timestamp column.")

if st.button("Generate Schedule", type="primary"):
    if not positions_file or not responses_file:
        st.error("Please upload both CSV files.")
        st.stop()

    positions_df = read_csv_robust(positions_file, "positions")
    responses_df = read_csv_robust(responses_file, "responses")

    # Detect name columns
    try:
        name_col_positions = positions_df.columns[0]
    except Exception as e:
        st.error(f"Could not detect a name column in positions CSV: {e}")
        st.stop()
    try:
        name_col_responses = detect_name_column(responses_df, fallback_first=False)
    except Exception as e:
        st.error(f"Could not detect a name column in responses CSV: {e}")
        st.stop()

    # Codes column (second column)
    codes_col = positions_df.columns[1] if positions_df.shape[1] >= 2 else None

    # Ensure name columns are strings
    positions_df[name_col_positions] = positions_df[name_col_positions].astype(str)
    responses_df[name_col_responses] = responses_df[name_col_responses].astype(str)

    # Display name map (prefer Positions casing), plus normalized keys
    name_display_map = build_display_name_map(positions_df, name_col_positions)

    # Role columns (from third column onward) — define all locations to schedule
    raw_role_cols = [c for c in positions_df.columns[2:] if is_priority_col(positions_df[c])]
    # Exclude plain "Entrance greeter"
    excluded_norm = {normalize(x) for x in EXCLUDED_ROLES}
    role_cols = [c for c in raw_role_cols if normalize(strip_capacity_tag(c)) not in excluded_norm]
    if not role_cols:
        st.error("No usable role columns detected in positions CSV (from the third column onwards).")
        st.stop()

    # Build eligibility (+ role code flags) — keys are normalized names
    long_df, role_codes = build_long_df(positions_df, name_col_positions, role_cols, codes_col=codes_col)
    if long_df.empty:
        st.error("No eligible assignments found (after applying Director=1-only rule and removing 0s).")
        st.stop()

    # Parse month & dates
    try:
        year, month, date_map, service_dates, sheet_name = parse_month_and_dates_from_headers(responses_df)
    except Exception as e:
        st.error(f"Could not parse month & dates from responses: {e}")
        st.stop()

    # Availability (latest response per normalized name)
    availability, service_dates, few_yes_list_norm, display_from_responses = parse_availability(responses_df, name_col_responses, date_map)

    # Update display map with response names only for keys missing from Positions
    for k, disp in display_from_responses.items():
        name_display_map.setdefault(k, disp)

    # MAIN scheduling (uses normalized keys)
    schedule_cells, assign_count_norm, slot_rows, slot_to_role, eligibility, people_norm, p1_roles_by_person = main_pass_schedule(
        long_df, availability, service_dates, role_codes, all_role_headers=role_cols
    )

    # Build schedule table (render display names)
    schedule_df = build_schedule_df(schedule_cells, slot_rows, service_dates, name_display_map)

    # Stats
    total_slots = schedule_df.size
    filled_slots = int((schedule_df != "").sum().sum())
    fill_rate = (filled_slots / total_slots) if total_slots else 0.0
    unfilled = total_slots - filled_slots

    # Per-person summary (+ details)
    per_series = pd.Series(assign_count_norm, name="Assignments")
    # map normalized index to display names
    per_series.index = [name_display_map.get(k, k) for k in per_series.index]
    per_person = (
        per_series.sort_values(ascending=False)
        .reset_index()
        .rename(columns={"index": "Person"})
    )
    details_lookup = build_person_assignment_details(schedule_cells, name_display_map)
    per_person["Locations & Dates"] = per_person["Person"].map(lambda nm: details_lookup.get(nm, ""))

    # <2 yes list — convert to display
    few_yes_display = [name_display_map.get(k, k) for k in few_yes_list_norm]

    # Unmet P1 list — compute with normalized keys, then display
    served_p1_people_norm = served_in_priority_one(schedule_cells, p1_roles_by_person, slot_to_role)
    p1_people_norm = sorted([p for p in people_norm if p1_roles_by_person.get(p)])
    unmet_p1_norm = [p for p in p1_people_norm if p not in served_p1_people_norm]
    unmet_p1_display = [name_display_map.get(k, k) for k in unmet_p1_norm]

    st.success(f"Schedule generated for **{sheet_name}**")
    st.write(f"Filled slots: **{filled_slots} / {total_slots}**  (Fill rate: **{fill_rate:.1%}**)  •  Unfilled: **{unfilled}**")

    st.subheader("Schedule (each slot is its own row)")
    st.dataframe(schedule_df, use_container_width=True)

    st.subheader("Assignment Summary")
    st.dataframe(per_person, use_container_width=True)

    st.subheader("People with < 2 'Yes' dates (for reference only)")
    few_yes_df = pd.DataFrame({"Person": few_yes_display})
    st.dataframe(few_yes_df, use_container_width=True)

    if unmet_p1_display:
        st.subheader("Unmet Priority-1 requirement (info)")
        st.caption("These volunteers have at least one Priority-1 location but were not scheduled into any Priority-1 slot (capacity/availability constraints).")
        st.dataframe(pd.DataFrame({"Person": unmet_p1_display}), use_container_width=True)

    # Excel export
    wb = Workbook()
    ws = wb.create_sheet(sheet_name) if wb.active.title != "Sheet" else wb.active
    ws.title = sheet_name

    header = ["Position / Slot"] + [d.strftime("%Y-%m-%d") for d in service_dates]
    ws.append(header)
    for row_name in slot_rows:
        # render display names
        row_vals = [row_name] + [
            ", ".join([name_display_map.get(nm, nm) for nm in schedule_cells[(row_name, d)]])
            for d in service_dates
        ]
        ws.append(row_vals)
    excel_autofit(ws)

    ws3 = wb.create_sheet("Assignment Summary")
    ws3.append(["Person", "Assignments", "Locations & Dates"])
    for _, r in per_person.iterrows():
        ws3.append([r["Person"], int(r["Assignments"]), r.get("Locations & Dates", "")])
    excel_autofit(ws3)

    if few_yes_display:
        ws2 = wb.create_sheet("Fewer than 2 Yes (info)")
        ws2.append(["Person"])
        for p in few_yes_display:
            ws2.append([p])
        excel_autofit(ws2)

    if unmet_p1_display:
        ws4 = wb.create_sheet("Unmet Priority-1 (info)")
        ws4.append(["Person"])
        for p in unmet_p1_display:
            ws4.append([p])
        excel_autofit(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    st.download_button(
        "Download Excel (.xlsx)",
        data=buf,
        file_name=f"uKids_schedule_{sheet_name.replace(' ','_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.info("Upload the two CSV files (any names), then click **Generate Schedule**.")
