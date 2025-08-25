# ukids_scheduler_app.py
# Streamlit uKids scheduler with:
# - Priority-1 pre-pass (each P1-eligible volunteer gets at least one P1 slot if possible)
# - Most-recent availability per volunteer (by Timestamp)
# - Slot plan built dynamically from all roles in Positions CSV headers
# - Eligibility enforcement (0 = not allowed)
# - Only volunteers present in BOTH files are considered
# - Single "Age 9 classroom" and single-person-per-slot guard
# - Assignment Summary includes "Locations & Dates"
# - Excel export with all sheets

import io
import re
import base64
from collections import defaultdict, Counter
from datetime import datetime, date

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="uKids Scheduler", layout="wide")
st.title("uKids Scheduler")

st.markdown(
    """
    <style>
      .stApp { background: #000; color: #fff; }
      .stButton>button, .stDownloadButton>button { background:#444; color:#fff; }
      .stDataFrame { background:#111; }
      .stAlert { color:#111; }
    </style>
    """,
    unsafe_allow_html=True,
)

# Optional logo (ignore if missing)
for logo_name in ["image(1).png", "image.png", "logo.png"]:
    try:
        with open(logo_name, "rb") as img_file:
            encoded = base64.b64encode(img_file.read()).decode()
            st.markdown(
                f"<div style='text-align:center'><img src='data:image/png;base64,{encoded}' width='520'></div>",
                unsafe_allow_html=True,
            )
            break
    except Exception:
        pass

# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────
MONTH_ALIASES = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}
YES_SET = {"yes", "y", "true", "available"}

def read_csv_robust(uploaded_file, label_for_error):
    """Read a Streamlit UploadedFile into a DataFrame, trying multiple encodings and separators."""
    raw = uploaded_file.getvalue()
    encodings = ["utf-8", "utf-8-sig", "cp1252", "iso-8859-1"]
    seps = [None, ",", ";", "\t", "|"]

    last_err = None
    for enc in encodings:
        for sep in seps:
            try:
                df = pd.read_csv(io.BytesIO(raw), encoding=enc, engine="python", sep=sep)
                if df.shape[1] == 0:
                    raise ValueError("Parsed 0 columns.")
                return df
            except Exception as e:
                last_err = f"{type(e).__name__}: {e}"
                continue
    st.error(
        f"Could not read {label_for_error} CSV. Last error: {last_err}. "
        "Try re-exporting as CSV (UTF-8) or remove unusual characters in headers."
    )
    st.stop()

def detect_name_column(df: pd.DataFrame, fallback_first: bool = True) -> str:
    candidates = [
        "What is your name AND surname?",
        "What is your name and surname?",
        "Name",
        "Full name",
        "Full names",
    ]
    cols_l = {str(c).strip().lower(): c for c in df.columns}
    for c in candidates:
        key = c.strip().lower()
        if key in cols_l:
            return cols_l[key]
    for c in df.columns:
        if isinstance(c, str) and "name" in c.lower():
            return c
    if fallback_first:
        return df.columns[0]
    raise ValueError("Could not detect a 'name' column.")

def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s).lower()).strip()

def is_priority_col(series: pd.Series) -> bool:
    vals = pd.to_numeric(series, errors="coerce").dropna()
    if len(vals) == 0:
        return False
    return (vals.min() >= 0) and (vals.max() <= 5)

def parse_month_and_dates_from_headers(responses_df: pd.DataFrame):
    avail_cols = [c for c in responses_df.columns if isinstance(c, str) and c.strip().lower().startswith("are you available")]
    if not avail_cols:
        avail_cols = [
            c for c in responses_df.columns
            if isinstance(c, str)
            and re.search(r"\b\d{1,2}\b", c.lower())
            and any(m in c.lower() for m in MONTH_ALIASES)
        ]
    if not avail_cols:
        raise ValueError("No availability columns found. Expect headings like 'Are you available 7 September?'")

    info = []
    for c in avail_cols:
        low = c.lower()
        mname = None
        for alias in MONTH_ALIASES:
            if alias in low:
                mname = alias
                break
        day_m = re.search(r"\b(\d{1,2})\b", low)
        if mname and day_m:
            info.append((c, MONTH_ALIASES[mname], int(day_m.group(1))))
    if not info:
        raise ValueError("Could not parse day/month from availability headers.")

    months = {m for _, m, _ in info}
    if len(months) > 1:
        raise ValueError(f"Multiple months detected in availability headers: {sorted(months)}. Upload one month at a time.")
    month = months.pop()

    if "Timestamp" in responses_df.columns:
        years = pd.to_datetime(responses_df["Timestamp"], errors="coerce").dt.year.dropna().astype(int)
        year = int(years.mode().iloc[0]) if not years.empty else date.today().year
    else:
        year = date.today().year

    date_map = {c: pd.Timestamp(datetime(year, month, d)).normalize() for c, _, d in info}
    service_dates = sorted(set(date_map.values()))
    sheet_name = f"{pd.Timestamp(year=year, month=month, day=1):%B %Y}"
    return year, month, date_map, service_dates, sheet_name

def build_long_df(people_df: pd.DataFrame, name_col: str, role_cols, codes_col: str = None):
    """
    returns:
      - long_df with columns [person, role, priority]
      - role_codes flags per person: has_D / has_BL / has_PL / has_EL / has_SL + raw_codes
    """
    records = []
    role_codes = {}
    for _, r in people_df.iterrows():
        person = str(r[name_col]).strip()
        if not person or person.lower() == "nan":
            continue

        # parse codes (2nd column)
        flags = {"has_D": False, "has_BL": False, "has_PL": False, "has_EL": False, "has_SL": False, "raw": ""}
        if codes_col and codes_col in people_df.columns:
            raw = str(r.get(codes_col, "")).strip()
            flags["raw"] = raw
            toks = re.findall(r"[A-Za-z]+", raw.upper())
            for t in toks:
                if t == "D":
                    flags["has_D"] = True
                elif t == "BL":
                    flags["has_BL"] = True
                elif t == "PL":
                    flags["has_PL"] = True
                elif t == "EL":
                    flags["has_EL"] = True
                elif t == "SL":
                    flags["has_SL"] = True
        role_codes[person] = flags

        for role in role_cols:
            pr = pd.to_numeric(r[role], errors="coerce")
            if pd.isna(pr):
                continue
            pr = int(round(pr))
            if pr >= 1:
                # If Director: only eligible for priority==1 roles
                if flags["has_D"] and pr != 1:
                    continue
                records.append({"person": person, "role": role, "priority": pr})
    return pd.DataFrame(records), role_codes

def dedupe_latest_by_name(df: pd.DataFrame, name_col: str) -> pd.DataFrame:
    """Keep only the most recent response per person using the 'Timestamp' column if present."""
    if "Timestamp" in df.columns:
        ts = pd.to_datetime(df["Timestamp"], errors="coerce")
        df2 = df.assign(_ts=ts).sort_values("_ts")
        latest = df2.groupby(name_col, as_index=False).tail(1).drop(columns=["_ts"])
        return latest
    return df.groupby(name_col, as_index=False).tail(1)

def parse_availability(responses_df: pd.DataFrame, name_col_resp: str, date_map):
    # Use only the latest response per person
    responses_latest = dedupe_latest_by_name(responses_df, name_col_resp)

    availability = {}
    yes_counts = Counter()
    for _, row in responses_latest.iterrows():
        nm = str(row.get(name_col_resp, "")).strip()
        if not nm or nm.lower() == "nan":
            continue
        availability.setdefault(nm, {})
        for col, dt in date_map.items():
            ans = str(row.get(col, "")).strip().lower()
            is_yes = ans in YES_SET
            availability[nm][dt] = is_yes
            if is_yes:
                yes_counts[nm] += 1
    few_yes = sorted([n for n, c in yes_counts.items() if c < 2])
    service_dates = sorted(set(date_map.values()))
    return availability, service_dates, few_yes

# ──────────────────────────────────────────────────────────────────────────────
# Dynamic slot plan — built from Positions CSV headers (role_cols)
# Known roles keep custom capacities; others default to 1
# ──────────────────────────────────────────────────────────────────────────────
DEFAULT_CAPS = {
    # Age 1
    "age 1 leader": 1,
    "age 1 classroom": 5,
    "age 1 nappies": 1,
    "age 1 bags girls": 1,
    "age 1 bags boys": 1,
    # Age 2
    "age 2 leader": 1,
    "age 2 classroom": 4,
    "age 2 nappies": 1,
    "age 2 bags girls": 1,
    "age 2 bags boys": 1,
    # Age 3
    "age 3 leader": 1,
    "age 3 classroom": 4,
    "age 3 bags": 1,
    # Age 4
    "age 4 leader": 1,
    "age 4 classroom": 4,
    # Age 5
    "age 5 leader": 1,
    "age 5 classroom": 3,
    # Age 6
    "age 6 leader": 1,
    "age 6 classroom": 3,
    # Age 7
    "age 7 leader": 1,
    "age 7 classroom": 2,
    # Age 8
    "age 8 leader": 1,
    "age 8 classroom": 2,
    # Age 9
    "age 9 leader": 1,
    "age 9 classroom": 1,
    # Age 10
    "age 10 leader": 1,
    "age 10 classroom": 1,
    # Age 11
    "age 11 leader": 1,
    "age 11 classroom": 1,
    # Special Needs
    "special needs leader": 1,
    "special needs classroom": 2,
    # Common extras
    "info": 4,  # <-- this is why you saw Info #1..#4
    "ukids setup": 4,
    "outside assistant": 2,
    "helping ninja and check in leader": 1,  # requires ukids leader
    "helping ninja": 2,
    "ukids hall": 4,
    # Brooklyn capacities (updated per your request)
    "brooklyn runner": 2,
    "brooklyn babies leader": 1,          # 1 leader
    "brooklyn babies serving girl": 2,    # +2 serving girls = total 3
    "brooklyn preschool leader": 1,       # 1 leader
    "brooklyn preschool": 3,              # +3 = total 4
}

def build_slot_plan_dynamic(all_role_headers):
    slot_plan = {}
    norm_caps = {normalize(k): v for k, v in DEFAULT_CAPS.items()}
    for role in all_role_headers:
        cap = norm_caps.get(normalize(role), 1)  # default to 1 when unknown
        try:
            cap_i = int(cap)
        except Exception:
            cap_i = 1
        slot_plan[role] = max(1, cap_i)
    return slot_plan

def expand_roles_to_slots(slot_plan):
    slot_rows = []
    slot_index = {}
    for role, n in slot_plan.items():
        if n <= 0:
            continue
        if n == 1:
            lab = role
            slot_rows.append(lab)
            slot_index[lab] = role
        else:
            for i in range(1, n + 1):
                lab = f"{role} #{i}"
                slot_rows.append(lab)
                slot_index[lab] = role
    return slot_rows, slot_index

def compute_p1_roles_by_person(long_df):
    """Return {person: set of base roles with priority == 1}."""
    p1 = defaultdict(set)
    for _, r in long_df.iterrows():
        if int(r["priority"]) == 1:
            p1[str(r["person"]).strip()].add(str(r["role"]).strip())
    return p1

def served_in_priority_one(assignments_by_cell, p1_roles_by_person, slot_to_role):
    """Find people assigned to any slot whose base role is in their P1 set."""
    served = set()
    for (row_name, _d), names in assignments_by_cell.items():
        base_role = slot_to_role.get(row_name, row_name)
        for nm in names:
            if base_role in p1_roles_by_person.get(nm, set()):
                served.add(nm)
    return served

def is_ukids_leader(flags: dict) -> bool:
    return bool(flags.get("has_BL") or flags.get("has_PL") or flags.get("has_EL") or flags.get("has_SL") or flags.get("has_D"))

def base_max_for_person(flags: dict) -> int:
    # Monthly caps (not per Sunday): Director=1, others=2
    return 1 if flags.get("has_D", False) else 2

# ──────────────────────────────────────────────────────────────────────────────
# Core scheduling with a P1 pre-pass (no special group passes)
# ──────────────────────────────────────────────────────────────────────────────
def build_eligibility(long_df: pd.DataFrame):
    elig = defaultdict(set)
    for _, r in long_df.iterrows():
        elig[str(r["person"]).strip()].add(str(r["role"]).strip())
    return elig

def role_allowed_for_person(eligibility, person, base_role):
    elig_roles = eligibility.get(person, set())
    if base_role in elig_roles:
        return True
    nb = normalize(base_role)
    for er in elig_roles:
        if normalize(er) == nb:
            return True
    return False

def main_pass_schedule(long_df, availability, service_dates, role_codes, all_role_headers):
    slot_plan = build_slot_plan_dynamic(all_role_headers)
    slot_rows, slot_to_role = expand_roles_to_slots(slot_plan)
    eligibility = build_eligibility(long_df)

    # Only schedule people present in both sources
    people = sorted(set(eligibility.keys()) & set(availability.keys()))

    # storage: {(row_name, date): [names]}
    schedule_cells = {(slot, d): [] for slot in slot_rows for d in service_dates}
    assign_count = defaultdict(int)

    def slot_sort_key(s):
        if "leader" in s.lower():
            return (0, s)
        if "classroom" in s.lower():
            return (1, s)
        return (2, s)
    slot_rows_sorted = sorted(slot_rows, key=slot_sort_key)

    # -------- P1 PRE-PASS: give each P1-eligible volunteer one P1 slot if possible
    p1_roles_by_person = compute_p1_roles_by_person(long_df)
    avail_count = {p: sum(1 for d in service_dates if availability.get(p, {}).get(d, False)) for p in people}
    people_order = sorted([p for p in people if p1_roles_by_person.get(p)], key=lambda p: (avail_count.get(p, 0), p.lower()))

    for p in people_order:
        flags = role_codes.get(p, {})
        monthly_cap = base_max_for_person(flags)
        if assign_count[p] >= monthly_cap:
            continue

        got_one = False
        for d in service_dates:  # try earliest dates first
            if not availability.get(p, {}).get(d, False):
                continue
            # avoid double-booking within the day
            assigned_today = {nm for (rn, dd), names in schedule_cells.items() if dd == d for nm in names}
            if p in assigned_today:
                continue

            # find an open slot whose base role is in their P1 set
            for slot_row in slot_rows_sorted:
                base_role = slot_to_role[slot_row]
                if base_role not in p1_roles_by_person[p]:
                    continue
                if not role_allowed_for_person(eligibility, p, base_role):
                    continue
                # special gate: "Helping ninja and check in leader" requires a uKids leader
                if normalize(base_role) == normalize("Helping ninja and check in leader") and not is_ukids_leader(flags):
                    continue
                # slot available? (single person per slot)
                if len(schedule_cells[(slot_row, d)]) == 0:
                    schedule_cells[(slot_row, d)].append(p)
                    assign_count[p] += 1
                    got_one = True
                    break
            if got_one:
                break

    # -------- GENERAL MAIN PASS
    for d in service_dates:
        assigned_today = set(nm for (rn, dd), names in schedule_cells.items() if dd == d for nm in names)
        for slot_row in slot_rows_sorted:
            base_role = slot_to_role[slot_row]

            # if this slot is already filled by the P1 pre-pass, skip it
            if len(schedule_cells[(slot_row, d)]) >= 1:
                continue

            # build candidate list
            cands = []
            for p in people:
                flags = role_codes.get(p, {})
                if assign_count[p] >= base_max_for_person(flags):
                    continue
                if p in assigned_today:
                    continue
                if not availability.get(p, {}).get(d, False):
                    continue
                if not role_allowed_for_person(eligibility, p, base_role):
                    continue
                # special gate: "Helping ninja and check in leader" requires a uKids leader
                if normalize(base_role) == normalize("Helping ninja and check in leader") and not is_ukids_leader(flags):
                    continue
                cands.append(p)

            if cands:
                cands.sort(key=lambda name: assign_count[name])
                chosen = cands[0]
                schedule_cells[(slot_row, d)].append(chosen)
                assign_count[chosen] += 1
                assigned_today.add(chosen)

    return schedule_cells, assign_count, slot_rows, slot_to_role, eligibility, people, p1_roles_by_person

# ──────────────────────────────────────────────────────────────────────────────
# Export & UI helpers
# ──────────────────────────────────────────────────────────────────────────────
def excel_autofit(ws):
    for col_idx, column_cells in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1
    ):
        max_len = 0
        for cell in column_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 80)

def build_schedule_df(schedule_cells, slot_rows, service_dates):
    cols = [d.strftime("%Y-%m-%d") for d in service_dates]
    schedule_df = pd.DataFrame(index=slot_rows, columns=cols)
    for (slot_row, d), names in schedule_cells.items():
        schedule_df.loc[slot_row, d.strftime("%Y-%m-%d")] = ", ".join(names)
    return schedule_df.fillna("")

def build_person_assignment_details(schedule_cells):
    per = defaultdict(list)
    for (slot_row, d), names in schedule_cells.items():
        for nm in names:
            per[nm].append((d, slot_row))
    details = {}
    for nm, items in per.items():
        items_sorted = sorted(items, key=lambda x: x[0])
        details[nm] = "; ".join([f"{dt.strftime('%Y-%m-%d')} — {slot}" for dt, slot in items_sorted])
    return details

# ──────────────────────────────────────────────────────────────────────────────
# UI
# ──────────────────────────────────────────────────────────────────────────────
st.subheader("1) Upload files (CSV — any filename)")
c1, c2 = st.columns(2)
with c1:
    positions_file = st.file_uploader("Serving positions (CSV)", type=["csv"], key="positions_csv_any")
with c2:
    responses_file = st.file_uploader("Availability responses (CSV)", type=["csv"], key="responses_csv_any")

st.caption("• Positions CSV: first col = volunteer names; second col = role codes (e.g., D/BL/PL/EL/SL), other cols = roles with values 0–5 (0 = not allowed).")
st.caption("• Responses CSV: includes a name column and columns like 'Are you available 7 September?' plus a Timestamp column.")

if st.button("Generate Schedule", type="primary"):
    if not positions_file or not responses_file:
        st.error("Please upload both CSV files.")
        st.stop()

    positions_df = read_csv_robust(positions_file, "positions")
    responses_df = read_csv_robust(responses_file, "responses")

    # Detect name columns
    try:
        name_col_positions = positions_df.columns[0]  # first column is full names
    except Exception as e:
        st.error(f"Could not detect a name column in positions CSV: {e}")
        st.stop()

    try:
        name_col_responses = detect_name_column(responses_df, fallback_first=False)
    except Exception as e:
        st.error(f"Could not detect a name column in responses CSV: {e}")
        st.stop()

    # Codes column (second column)
    codes_col = None
    if positions_df.shape[1] >= 2:
        codes_col = positions_df.columns[1]

    # Ensure name columns are strings
    positions_df[name_col_positions] = positions_df[name_col_positions].astype(str)
    responses_df[name_col_responses] = responses_df[name_col_responses].astype(str)

    # Role columns (from third column onward) — these define ALL locations to schedule
    role_cols = [c for c in positions_df.columns[2:] if is_priority_col(positions_df[c])]
    if not role_cols:
        st.error("No role columns with priorities (0..5) detected in the positions CSV (from third column onwards).")
        st.stop()

    # Build eligibility (+ role code flags)
    long_df, role_codes = build_long_df(positions_df, name_col_positions, role_cols, codes_col=codes_col)
    if long_df.empty:
        st.error("No eligible assignments found (after applying Director=1-only rule and removing 0s).")
        st.stop()

    # Parse dates
    try:
        year, month, date_map, service_dates, sheet_name = parse_month_and_dates_from_headers(responses_df)
    except Exception as e:
        st.error(f"Could not parse month & dates from responses: {e}")
        st.stop()

    # Availability (dedup to latest response) & <2-Yes list (INFO ONLY)
    availability, service_dates, few_yes_list = parse_availability(responses_df, name_col_responses, date_map)

    # MAIN (with P1 pre-pass) — includes single-person-per-slot guard
    schedule_cells, assign_count, slot_rows, slot_to_role, eligibility, people, p1_roles_by_person = main_pass_schedule(
        long_df, availability, service_dates, role_codes, all_role_headers=role_cols
    )

    # Build schedule DataFrame
    schedule_df = build_schedule_df(schedule_cells, slot_rows, service_dates)

    # Stats
    total_slots = schedule_df.size
    filled_slots = int((schedule_df != "").sum().sum())
    fill_rate = (filled_slots / total_slots) if total_slots else 0.0
    unfilled = total_slots - filled_slots

    # Build per-person summary, with details column
    per_person = (
        pd.Series(assign_count, name="Assignments")
        .sort_values(ascending=False)
        .reset_index()
        .rename(columns={"index": "Person"})
    )
    details_lookup = build_person_assignment_details(schedule_cells)
    per_person["Locations & Dates"] = per_person["Person"].map(lambda nm: details_lookup.get(nm, ""))

    # Compute unmet P1 requirement list (people who have P1 roles but never got a P1 assignment)
    served_p1_people = served_in_priority_one(schedule_cells, p1_roles_by_person, slot_to_role)
    p1_people = sorted([p for p in people if p1_roles_by_person.get(p)])
    unmet_p1 = [p for p in p1_people if p not in served_p1_people]

    st.success(f"Schedule generated for **{sheet_name}**")
    st.write(f"Filled slots: **{filled_slots} / {total_slots}**  (Fill rate: **{fill_rate:.1%}**)  •  Unfilled: **{unfilled}**")

    st.subheader("Schedule (each slot is its own row)")
    st.dataframe(schedule_df, use_container_width=True)

    st.subheader("Assignment Summary")
    st.dataframe(per_person, use_container_width=True)

    st.subheader("People with < 2 'Yes' dates (for reference only)")
    few_yes_df = pd.DataFrame({"Person": few_yes_list})
    st.dataframe(few_yes_df, use_container_width=True)

    if unmet_p1:
        st.subheader("Unmet Priority-1 requirement (info)")
        st.caption("These volunteers have at least one Priority-1 location but were not scheduled into any Priority-1 slot (capacity/availability constraints).")
        st.dataframe(pd.DataFrame({"Person": unmet_p1}), use_container_width=True)

    # Excel output
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header = ["Position / Slot"] + [d.strftime("%Y-%m-%d") for d in service_dates]
    ws.append(header)
    for row_name in slot_rows:
        row_vals = [row_name] + [", ".join(schedule_cells[(row_name, d)]) for d in service_dates]
        ws.append(row_vals)
    excel_autofit(ws)

    # Assignment Summary (with details)
    ws3 = wb.create_sheet("Assignment Summary")
    ws3.append(["Person", "Assignments", "Locations & Dates"])
    for _, r in per_person.iterrows():
        ws3.append([r["Person"], int(r["Assignments"]), r.get("Locations & Dates", "")])
    excel_autofit(ws3)

    # Fewer than 2 Yes (optional sheet)
    if few_yes_list:
        ws2 = wb.create_sheet("Fewer than 2 Yes (info)")
        ws2.append(["Person"])
        for p in few_yes_list:
            ws2.append([p])
        excel_autofit(ws2)

    # Unmet P1 (optional sheet)
    if unmet_p1:
        ws4 = wb.create_sheet("Unmet Priority-1 (info)")
        ws4.append(["Person"])
        for p in unmet_p1:
            ws4.append([p])
        excel_autofit(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    st.download_button(
        "Download Excel (.xlsx)",
        data=buf,
        file_name=f"uKids_schedule_{sheet_name.replace(' ','_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.info("Upload the two CSV files (any names), then click **Generate Schedule**.")
